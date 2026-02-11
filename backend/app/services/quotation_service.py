"""
Quotation Service - Generate Excel quotations from template
"""
from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
import tempfile
import os
import json
from datetime import datetime


PAYMENT_TERMS_OPTIONS = [
    "Option 1: 50% advance upon signing, 50% upon final delivery",
    "Option 2: 50% advance, 30% at dry port, 20% upon delivery",
    "Option 3: 60% advance upon signing, 40% upon final delivery",
    "Option 4: 30% advance, 40% at dry port, 30% upon delivery",
    "Option 5: 50% advance, 25% at dry port, 25% upon delivery",
]


def get_next_invoice_number() -> str:
    """
    Get the next sequential invoice number starting from 001.
    Stores the counter in a JSON file.
    """
    # Get the backend directory
    backend_dir = Path(__file__).parent.parent.parent
    counter_file = backend_dir / "invoice_counter.json"
    
    # Read current counter
    if counter_file.exists():
        try:
            with open(counter_file, 'r') as f:
                data = json.load(f)
                current_number = data.get('last_invoice_number', 0)
        except (json.JSONDecodeError, KeyError):
            current_number = 0
    else:
        current_number = 0
    
    # Increment
    next_number = current_number + 1
    
    # Save back
    try:
        with open(counter_file, 'w') as f:
            json.dump({'last_invoice_number': next_number}, f)
    except Exception:
        # If file write fails, continue anyway
        pass
    
    # Format as 001, 002, etc.
    return f"{next_number:03d}"


def generate_quotation(
    template_path: str,
    output_path: Optional[str] = None,
    products: Optional[List[Dict[str, Any]]] = None,
    payment_option: int = 1,
    form_type: str = "Baracoda",
    company_name: Optional[str] = None,
) -> str:
    """
    Generate a professional quotation Excel file from template.
    
    Args:
        template_path: Path to your styled Excel template
        output_path: Where to save the generated file (if None, creates temp file)
        products: List of product dicts with 'product_name', 'vendor_name', 'unit_price', 'quantity'
        payment_option: 1-5 corresponding to payment term options
        form_type: Type of quotation form (Baracoda, Nyumb-Chem, Bet-chem)
    
    Returns:
        Path to the generated Excel file
    """
    # Load template WITH formatting preserved
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    wb = load_workbook(template_path)
    ws = wb.active

    # ===== STEP 0: Determine template structure based on form type =====
    # Baracoda: Products B19-E23, Subtotal B26, Total E26
    # Bet-chem: Products C17-E21, Subtotal E22, Total E24
    # Nyumb-Chem: Products B10-E12, Quantity C10-C12, Price D10-D12, Total E10-E12, Subtotal E16, Total E17
    if form_type == "Nyumb-Chem":
        # Nyumb-Chem structure (based on actual template):
        # Row 9: Headers (A9=ITEM DESCRIPTION, B9=UNIT, C9=UNIT PRICE, D9=QTY, E9=TOTAL)
        # A10, A11... = Product Description
        # B10, B11... = UNIT (can be blank/0)
        # C10, C11... = Unit Price
        # D10, D11... = Quantity
        # E10, E11... = Total (Unit Price × Quantity) = C × D
        # D16 = "SubTotal" label, E16 = Subtotal (sum of E column)
        # D17 = "Total" label, E17 = Total (Subtotal × 1.15)
        # A25 = "Term of payment" label, B25 = Payment Terms value
        product_start_row = 10
        product_col_desc = "A"  # Description in column A
        product_col_unit = "B"  # UNIT column (can be blank)
        product_col_unit_price = "C"  # Unit Price in column C
        product_col_qty = "D"   # Quantity in column D
        product_col_total = "E"  # Total in column E (C × D)
        # Base positions: D16/E16 (subtotal), D17/E17 (total)
        base_subtotal_row = 16
        base_vat_row = 16  # Not used separately for Nyumb-Chem
        base_total_row = 17
        subtotal_label_col = "D"  # "SubTotal" label in D column
        subtotal_col = "E"  # Subtotal value in E column
        total_label_col = "D"  # "Total" label in D column
        total_col = "E"  # Total value in E column
        company_name_row = 7
        company_name_col = "D"  # Company name in D7 (CONSIGNEE/IMPORTER section)
        page_break_row = 24  # Before terms (terms at row 25)
        payment_terms_label_row = 25  # "Term of payment" label in A25
        payment_terms_row = 25  # Payment terms value in B25
        payment_terms_label_col = "A"  # Label in column A
        payment_terms_col = "B"  # Payment terms value in column B
    elif form_type == "Bet-chem":
        # Bet-chem structure:
        # C17, C18... = Product Description
        # D17, D18... = Quantity
        # E17, E18... = Unit Price
        # F17, F18... = Total (Quantity × Unit Price)
        # E22 = Subtotal (sum of F column)
        # E24 = Grand Total (Subtotal × 1.15)
        product_start_row = 17
        product_col_desc = "C"  # Description
        product_col_qty = "D"   # Quantity
        product_col_unit_price = "E"  # Unit Price
        product_col_total = "F"  # Total (Quantity × Unit Price)
        # Subtotal and total rows are dynamic - will be calculated after products are added
        # Base positions: F22 (subtotal), F23 (VAT), F24 (total) - moved one column right
        base_subtotal_row = 22
        base_vat_row = 23
        base_total_row = 24
        subtotal_col = "F"  # Subtotal in F column (moved from E)
        vat_col = "F"  # VAT rate in F column (moved from E)
        total_col = "F"  # Grand Total in F column (moved from E)
        company_name_row = 11
        company_name_col = "C"  # Bet-chem uses column C for address/company
        page_break_row = 25  # Before bank details
        payment_terms_header_row = 32  # "Terms and Conditions" header in C32
        payment_terms_row = 33  # Payment terms value in C33 (directly under header)
        payment_terms_col = "C"  # Payment terms in column C
    else:
        # Baracoda structure (default)
        product_start_row = 19
        product_col_desc = "B"  # Description
        product_col_qty = "D"   # Quantity
        product_col_price = "C"  # Unit Price
        # Base positions: B26 (subtotal), C26 (VAT), E26 (total)
        base_subtotal_row = 26
        base_vat_row = 26
        base_total_row = 26
        subtotal_row = 26
        subtotal_col = "B"
        vat_row = 26
        vat_col = "C"  # VAT rate in C26
        total_row = 26
        total_col = "E"
        company_name_row = 11
        company_name_col = "B"  # Baracoda uses column B
        page_break_row = 28  # Before trade terms
        payment_terms_header_row = 28  # "TRADE TERMS & CONDITIONS" header
        payment_terms_row = 30  # Payment terms value in B30
        payment_terms_col = "B"  # Payment terms in column B

    # ===== STEP 0.5: Set Invoice Number (sequential starting from 001) =====
    invoice_number = get_next_invoice_number()
    invoice_set = False  # Initialize for both form types
    
    # Set invoice number and date for Bet-chem and Nyumb-Chem, or try to find existing location for Baracoda
    if form_type == "Bet-chem":
        # Set date in D9
        current_date = datetime.now().strftime("%d/%m/%Y")
        ws["D9"] = f"DATE: {current_date}"
        # Set invoice number in D10
        ws["D10"] = f"Invoice No. {invoice_number}"
        invoice_set = True
    elif form_type == "Nyumb-Chem":
        # Nyumb-Chem: Invoice number in A3 ("Invoice no.: # "), Date in A4 ("Date :")
        # Only add invoice number, don't remove existing text
        try:
            # Set invoice number in A3 - preserve existing text, just add number after #
            invoice_cell = ws["A3"]
            current_value = str(invoice_cell.value or "")
            if "#" in current_value:
                # Replace only the part after # with the invoice number
                parts = current_value.split("#", 1)
                if len(parts) == 2:
                    # Keep everything before #, add invoice number after #
                    ws["A3"] = f"{parts[0]}#{invoice_number}"
                else:
                    # Just append invoice number after #
                    ws["A3"] = f"{current_value}{invoice_number}"
            elif "Invoice no.:" in current_value or "Invoice no" in current_value:
                # Add # and invoice number if not present
                ws["A3"] = f"{current_value} #{invoice_number}"
            else:
                # If no invoice text, add it
                ws["A3"] = f"{current_value} Invoice no.: #{invoice_number}"
            invoice_set = True
        except Exception:
            pass
        
        # Set current date in A4 - preserve existing text, just add date
        try:
            current_date = datetime.now().strftime("%d/%m/%Y")
            date_cell = ws["A4"]
            current_date_value = str(date_cell.value or "")
            if "Date" in current_date_value:
                # If "Date :" exists, replace everything after it with the date
                if "Date :" in current_date_value:
                    parts = current_date_value.split("Date :", 1)
                    ws["A4"] = f"{parts[0]}Date : {current_date}"
                else:
                    ws["A4"] = f"{current_date_value} {current_date}"
            else:
                # Add date if not present
                ws["A4"] = f"{current_date_value} Date : {current_date}"
        except Exception:
            pass
    else:
        # For Baracoda, try to find existing invoice number field
        def get_top_left_cell(cell_ref):
            """Get the top-left cell of a merged range if the cell is merged, otherwise return the cell itself"""
            cell = ws[cell_ref]
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Return the top-left cell of the merged range
                    top_left_row = merged_range.min_row
                    top_left_col = merged_range.min_col
                    return ws.cell(row=top_left_row, column=top_left_col)
            return cell
        
        invoice_locations = ["D7", "E7", "D8"]
        for cell_ref in invoice_locations:
            try:
                cell = get_top_left_cell(cell_ref)
                if cell.value and "INVOICE NUMBER" in str(cell.value).upper():
                    current_value = str(cell.value)
                    if "#" in current_value:
                        parts = current_value.split("#")
                        if len(parts) > 1:
                            cell.value = f"{parts[0]}#{invoice_number}"
                        else:
                            cell.value = f"{current_value} {invoice_number}"
                    else:
                        cell.value = f"{current_value} #{invoice_number}"
                    invoice_set = True
                    break
            except Exception:
                # Skip if there's an error accessing the cell
                continue
        
        if not invoice_set:
            try:
                # Try to set in D7, handling merged cells
                cell = get_top_left_cell("D7")
                cell.value = f"INVOICE NUMBER: #{invoice_number}"
                invoice_set = True
            except Exception:
                # If D7 fails, try E7
                try:
                    cell = get_top_left_cell("E7")
                    cell.value = f"INVOICE NUMBER: #{invoice_number}"
                    invoice_set = True
                except Exception:
                    pass
        
        # Set current date for Baracoda
        try:
            current_date = datetime.now().strftime("%d/%m/%Y")
            date_cell = get_top_left_cell("D8")
            if date_cell.value and "DATE:" in str(date_cell.value).upper():
                date_cell.value = f"DATE: {current_date}"
            else:
                date_cell.value = f"DATE: {current_date}"
        except Exception:
            # If setting date fails, continue without error
            pass

    # ===== STEP 0.6: Add company name =====
    if company_name:
        if form_type == "Bet-chem":
            # Bet-chem: Company name in C11 (address field)
            ws[f"{company_name_col}{company_name_row}"] = company_name
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: Company name in D7 (CONSIGNEE/IMPORTER section, after "COMPANY NAME:")
            # The template has "COMPANY NAME:" in D7, so we need to append or replace
            try:
                current_value = ws[f"{company_name_col}{company_name_row}"].value or ""
                if "COMPANY NAME:" in str(current_value):
                    # Replace the company name part
                    parts = str(current_value).split("COMPANY NAME:")
                    if len(parts) > 1:
                        ws[f"{company_name_col}{company_name_row}"] = f"{parts[0]}COMPANY NAME:\n{company_name}"
                    else:
                        ws[f"{company_name_col}{company_name_row}"] = f"{current_value}\n{company_name}"
                else:
                    ws[f"{company_name_col}{company_name_row}"] = f"COMPANY NAME:\n{company_name}"
            except Exception:
                # If setting fails, try direct assignment
                ws[f"{company_name_col}{company_name_row}"] = company_name
        else:
            # Baracoda: Company name next to "To:" in B11
            ws[f"{company_name_col}{company_name_row}"] = f"To: {company_name}"

    # ===== STEP 1: Prepare products list =====
    if not products:
        products = []
    
    start_row = product_start_row
    num_products = len(products)
    
    # Determine how many rows to clear
    # Calculate the last product row we'll use
    max_allowed_products = 50  # Safety limit
    actual_products = min(num_products, max_allowed_products)
    last_product_row_needed = start_row + actual_products - 1 if actual_products > 0 else start_row - 1
    
    # Clear rows based on actual products needed
    # Allow clearing up to 50 products worth of rows (well beyond 20 required)
    # Terms will be positioned dynamically after totals, so we can clear more rows safely
    buffer_rows = 3
    end_clear_row = last_product_row_needed + buffer_rows
    
    # For safety, don't clear beyond row 60 (leaves plenty of room for terms which will be positioned dynamically)
    max_safe_clear = 60
    end_clear_row = min(end_clear_row, max_safe_clear)
    
    # ===== STEP 2: Clear previous product entries =====
    # Helper function to safely clear a cell (handles merged cells)
    def safe_clear_cell(cell_ref):
        """Clear a cell value, handling merged cells by clearing the top-left cell"""
        try:
            cell = ws[cell_ref]
            # Check if this cell is part of a merged range
            for merged_range in ws.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    # Get the top-left cell of the merged range
                    top_left_row = merged_range.min_row
                    top_left_col = merged_range.min_col
                    cell = ws.cell(row=top_left_row, column=top_left_col)
                    break
            # Only clear if this is the top-left cell of a merge or not merged
            # Check if this cell is actually the top-left of any merged range
            is_top_left = True
            for merged_range in ws.merged_cells.ranges:
                if (merged_range.min_row == cell.row and 
                    merged_range.min_col == cell.column):
                    is_top_left = True
                    break
                elif cell.coordinate in merged_range:
                    is_top_left = False
                    break
            
            if is_top_left:
                cell.value = None
        except Exception:
            # If clearing fails, skip this cell
            pass
    
    # Clear only the product columns for rows we'll use (preserves other template content)
    for row in range(start_row, end_clear_row + 1):
        safe_clear_cell(f"{product_col_desc}{row}")
        if form_type == "Bet-chem":
            # Bet-chem: Clear D (quantity), E (unit price) and F (total)
            safe_clear_cell(f"{product_col_qty}{row}")
            safe_clear_cell(f"{product_col_unit_price}{row}")
            safe_clear_cell(f"{product_col_total}{row}")
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: Clear A (description), B (unit), C (unit price), D (quantity), E (total)
            safe_clear_cell(f"{product_col_desc}{row}")
            safe_clear_cell(f"{product_col_unit}{row}")
            safe_clear_cell(f"{product_col_unit_price}{row}")
            safe_clear_cell(f"{product_col_qty}{row}")
            safe_clear_cell(f"{product_col_total}{row}")
        else:
            # Baracoda: Clear B (description), C (unit price), D (quantity), E (amount)
            safe_clear_cell(f"{product_col_qty}{row}")
            safe_clear_cell(f"{product_col_price}{row}")
            safe_clear_cell(f"E{row}")

    # ===== STEP 3: Populate new products (supports at least 20 products) =====
    # We'll calculate the safe limit after we know where totals will be
    # For now, allow up to 50 products (which is well within template limits)
    # Terms will be positioned dynamically after totals, so we don't need to limit here
    
    for idx, product in enumerate(products):
        row = start_row + idx
        
        # Support at least 20 products, but allow more if needed
        # Only stop if we've exceeded a reasonable limit (50 products max to prevent issues)
        max_allowed_products = 50
        if idx >= max_allowed_products:
            break
        
        # If this row is beyond our initial clear range, clear it now (only product columns)
        if row > end_clear_row:
            safe_clear_cell(f"{product_col_desc}{row}")
            if form_type == "Bet-chem":
                safe_clear_cell(f"{product_col_qty}{row}")
                safe_clear_cell(f"{product_col_unit_price}{row}")
                safe_clear_cell(f"{product_col_total}{row}")
            elif form_type == "Nyumb-Chem":
                safe_clear_cell(f"{product_col_unit}{row}")
                safe_clear_cell(f"{product_col_unit_price}{row}")
                safe_clear_cell(f"{product_col_qty}{row}")
                safe_clear_cell(f"{product_col_total}{row}")
            else:
                safe_clear_cell(f"{product_col_qty}{row}")
                safe_clear_cell(f"{product_col_price}{row}")
                safe_clear_cell(f"E{row}")
        
        # Product name with vendor - set to black color
        product_name = product.get("product_name", "")
        vendor_name = product.get("vendor_name", "")
        full_name = f"{product_name}, {vendor_name}" if vendor_name else product_name
        
        # SIMPLIFIED: Direct assignment
        if form_type == "Nyumb-Chem":
            ws[f"A{row}"].value = full_name
            ws[f"A{row}"].font = Font(color="000000")
        else:
            product_desc_cell = ws[f"{product_col_desc}{row}"]
            product_desc_cell.value = full_name
            product_desc_cell.font = Font(color="000000")  # Black color
        
        # Unit Price and Quantity - order depends on form type
        unit_price = product.get("unit_price", 0)
        quantity = product.get("quantity", 0)
        
        if form_type == "Bet-chem":
            # Bet-chem: Quantity in D, Unit price in E, Total in F
            qty_cell = ws[f"{product_col_qty}{row}"]
            qty_cell.value = quantity
            qty_cell.number_format = "#,##0"
            qty_cell.font = Font(color="000000")  # Black color
            
            price_cell = ws[f"{product_col_unit_price}{row}"]
            price_cell.value = unit_price
            price_cell.number_format = "#,##0.00"
            price_cell.font = Font(color="000000")  # Black color
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: SIMPLIFIED - Just set the values directly
            # A=Description, B=UNIT (blank), C=Unit Price, D=Quantity, E=Total
            
            # Set UNIT column B to blank
            ws[f"B{row}"].value = ""
            
            # Unit price in column C
            ws[f"C{row}"].value = unit_price
            ws[f"C{row}"].number_format = "#,##0.00"
            ws[f"C{row}"].font = Font(color="000000")
            
            # Quantity in column D
            ws[f"D{row}"].value = quantity
            ws[f"D{row}"].number_format = "#,##0"
            ws[f"D{row}"].font = Font(color="000000")
        else:
            # Baracoda: Description in B, Unit price in C, Quantity in D, Total in E
            qty_cell = ws[f"{product_col_qty}{row}"]
            qty_cell.value = quantity
            qty_cell.number_format = "#,##0"
            qty_cell.font = Font(color="000000")  # Black color
            
            price_cell = ws[f"{product_col_price}{row}"]
            price_cell.value = unit_price
            price_cell.number_format = "#,##0.00"
            price_cell.font = Font(color="000000")  # Black color
        
        # Calculate Total - set to black color
        if form_type == "Bet-chem":
            # Bet-chem: Total = D (quantity) × E (unit price) in column F
            total_cell = ws[f"{product_col_total}{row}"]
            total_cell.value = f"=PRODUCT({product_col_qty}{row},{product_col_unit_price}{row})"
            total_cell.number_format = "#,##0.00"
            total_cell.font = Font(color="000000")  # Black color
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: Total = C × D in column E - SIMPLIFIED
            ws[f"E{row}"].value = f"=C{row}*D{row}"
            ws[f"E{row}"].number_format = "#,##0.00"
            ws[f"E{row}"].font = Font(color="000000")
        else:
            # Baracoda: Amount = C (unit price) * D (quantity) in column E
            total_cell = ws[f"E{row}"]
            total_cell.value = f"=PRODUCT(C{row},D{row})"
            total_cell.number_format = "#,##0.00"
            total_cell.font = Font(color="000000")  # Black color

    # ===== STEP 4: Calculate last product row and update formulas =====
    # Calculate where the last product actually is
    # Use the actual number of products processed (may be limited by max_allowed_products)
    actual_products_processed = min(num_products, max_allowed_products)
    last_product_row = start_row + actual_products_processed - 1 if actual_products_processed > 0 else start_row - 1
    
    # Calculate dynamic subtotal/total rows (right after last product)
    if form_type == "Bet-chem":
        # Bet-chem: Subtotal 2 rows after last product, or use base position
        subtotal_row = max(base_subtotal_row, last_product_row + 2)
        vat_row = subtotal_row + 1
        total_row = subtotal_row + 2
    elif form_type == "Nyumb-Chem":
        # Nyumb-Chem: Subtotal 2 rows after last product, or use base position (E16)
        # Total is 1 row after subtotal
        subtotal_row = max(base_subtotal_row, last_product_row + 2)
        total_row = subtotal_row + 1
        vat_row = subtotal_row  # Not used separately
    else:
        # Baracoda: Subtotal 3 rows after last product, or use base position
        subtotal_row = max(base_subtotal_row, last_product_row + 3)
        vat_row = subtotal_row  # Same row for Baracoda
        total_row = subtotal_row  # Same row for Baracoda
    
    if num_products > 0:
        if form_type == "Bet-chem":
            # Bet-chem: Subtotal = SUM of Total column (F17:F...)
            # Use all products (unlimited)
            if num_products == 1:
                sum_formula = f"={product_col_total}{start_row}"
            else:
                # Sum all product rows up to last_product_row
                sum_formula = f"=SUM({product_col_total}{start_row}:{product_col_total}{last_product_row})"
            
            # Set subtotal, VAT, and total at calculated rows
            # Headers moved to E column (one column before F), numbers in F column
            # Add "Subtotal" header in E column (before F)
            subtotal_header_col = "E"
            subtotal_header_cell = ws[f"{subtotal_header_col}{subtotal_row}"]
            subtotal_header_cell.value = "Subtotal"
            subtotal_header_cell.font = Font(bold=True)
            
            ws[f"{subtotal_col}{subtotal_row}"] = sum_formula
            ws[f"{subtotal_col}{subtotal_row}"].number_format = "#,##0.00"
            
            # Add "VAT" header in E column (before F)
            vat_header_col = "E"
            vat_header_cell = ws[f"{vat_header_col}{vat_row}"]
            vat_header_cell.value = "VAT"
            vat_header_cell.font = Font(bold=True)
            
            ws[f"{vat_col}{vat_row}"] = 0.15
            ws[f"{vat_col}{vat_row}"].number_format = "0.00"
            
            # Add "Total" header in E column (before F)
            total_header_col = "E"
            total_header_cell = ws[f"{total_header_col}{total_row}"]
            total_header_cell.value = "Total"
            total_header_cell.font = Font(bold=True)
            
            ws[f"{total_col}{total_row}"] = f"={subtotal_col}{subtotal_row}*1.15"
            ws[f"{total_col}{total_row}"].number_format = "#,##0.00"
            ws[f"{total_col}{total_row}"].font = Font(bold=True)
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: Subtotal = SUM of Total column (E10:E...)
            # Use all products (unlimited)
            if num_products == 1:
                sum_formula = f"={product_col_total}{start_row}"
            else:
                # Sum all product rows up to last_product_row
                sum_formula = f"=SUM({product_col_total}{start_row}:{product_col_total}{last_product_row})"
            
            # Clear old base positions (16, 17) if they differ from calculated positions
            if subtotal_row != base_subtotal_row:
                safe_clear_cell(f"{subtotal_label_col}{base_subtotal_row}")
                safe_clear_cell(f"{subtotal_col}{base_subtotal_row}")
            if total_row != base_total_row:
                safe_clear_cell(f"{total_label_col}{base_total_row}")
                safe_clear_cell(f"{total_col}{base_total_row}")
            
            # Clear existing values at calculated rows before setting formulas
            safe_clear_cell(f"{subtotal_col}{subtotal_row}")
            safe_clear_cell(f"{total_col}{total_row}")
            safe_clear_cell(f"{subtotal_label_col}{subtotal_row}")
            safe_clear_cell(f"{total_label_col}{total_row}")
            
            # SIMPLIFIED: Just set the values directly
            # SubTotal label and formula
            ws[f"D{subtotal_row}"].value = "SubTotal"
            ws[f"D{subtotal_row}"].font = Font(bold=True)
            ws[f"E{subtotal_row}"].value = sum_formula
            ws[f"E{subtotal_row}"].number_format = "#,##0.00"
            ws[f"E{subtotal_row}"].font = Font(bold=True, color="000000")
            
            # Total label and formula
            ws[f"D{total_row}"].value = "Total"
            ws[f"D{total_row}"].font = Font(bold=True)
            ws[f"E{total_row}"].value = f"=E{subtotal_row}*1.15"
            ws[f"E{total_row}"].number_format = "#,##0.00"
            ws[f"E{total_row}"].font = Font(bold=True, color="000000")
        else:
            # Baracoda: Subtotal = SUM of Amount column (E19:E...)
            # Use all products (unlimited)
            if num_products == 1:
                sum_formula = f"=E{start_row}"
            else:
                # Sum all product rows up to last_product_row
                sum_formula = f"=SUM(E{start_row}:E{last_product_row})"
            
            # Set subtotal and total at calculated rows
            # Add "Subtotal" header one column before (A column before B)
            subtotal_header_col = "A"
            subtotal_header_cell = ws[f"{subtotal_header_col}{subtotal_row}"]
            subtotal_header_cell.value = "Subtotal"
            subtotal_header_cell.font = Font(bold=True)
            
            ws[f"{subtotal_col}{subtotal_row}"] = sum_formula
            ws[f"{subtotal_col}{subtotal_row}"].number_format = "#,##0.00"
            
            # Add "Total" header one column before (D column before E)
            total_header_col = "D"
            total_header_cell = ws[f"{total_header_col}{total_row}"]
            total_header_cell.value = "Total"
            total_header_cell.font = Font(bold=True)
            
            ws[f"{total_col}{total_row}"] = f"={subtotal_col}{subtotal_row}*1.15"
            ws[f"{total_col}{total_row}"].number_format = "#,##0.00"
            ws[f"{total_col}{total_row}"].font = Font(bold=True)
    else:
        # No products - set to 0
        ws[f"{subtotal_col}{subtotal_row}"] = 0
        ws[f"{subtotal_col}{subtotal_row}"].number_format = "#,##0.00"
        ws[f"{total_col}{total_row}"] = 0
        ws[f"{total_col}{total_row}"].number_format = "#,##0.00"

    # ===== STEP 5: Handle pagination - Move Trade Terms/Bank Details to next page if products exceed limit =====
    # Baracoda: If products exceed row 21, move trade terms to page 2
    # Bet-chem: If products exceed row 20, move bank details to page 2
    # Nyumb-Chem: If products exceed row 15, move terms to page 2
    if form_type == "Bet-chem":
        max_product_row_before_page_break = 20
    elif form_type == "Nyumb-Chem":
        max_product_row_before_page_break = 15  # Terms at B25, so break before if products go beyond row 15
    else:
        max_product_row_before_page_break = 21
    
    # Use the actual number of products processed (already calculated in STEP 4)
    # Check if products or totals exceed the threshold for page break
    # For Nyumb-Chem, also check if total_row exceeds the threshold
    if form_type == "Nyumb-Chem":
        # For Nyumb-Chem, check if total_row (which includes subtotal) exceeds threshold
        # Terms are at row 25, so we want page break before if total_row >= 24
        needs_page_break = last_product_row > max_product_row_before_page_break or total_row >= 24
    else:
        needs_page_break = last_product_row > max_product_row_before_page_break
    
    if needs_page_break:
        # When products exceed row 21, we need to ensure trade terms (row 28+) appear ONLY on page 2
        # Strategy: Insert a hard page break at row 28 and ensure proper page setup
        try:
            from openpyxl.worksheet.pagebreak import Break
            
            # Initialize row_breaks if needed
            if not hasattr(ws, 'row_breaks') or ws.row_breaks is None:
                ws.row_breaks = []
            
            # Set page break before trade terms/bank details section
            # page_break_row was already defined earlier based on form_type
            break_row = page_break_row
            
            # Remove any existing breaks at this row and add our break
            ws.row_breaks = [br for br in ws.row_breaks if br.id != break_row]
            ws.row_breaks.append(Break(id=break_row))
            
            # Configure page setup to respect manual page breaks
            ws.page_setup.fitToHeight = False
            ws.page_setup.fitToWidth = True
            
            # Ensure page breaks are enabled
            if hasattr(ws, 'sheet_properties'):
                ws.sheet_properties.pageSetUpPr = None  # Reset to default
            
        except Exception as e:
            # If page break API fails, don't insert rows as it distorts the template
            # Page breaks should be handled by Excel's print settings
            pass

    # ===== STEP 6: Set Payment Terms (directly under header, dynamic positioning) =====
    if 1 <= payment_option <= 5:
        payment_text = PAYMENT_TERMS_OPTIONS[payment_option - 1]
        
        # Format payment terms text like the example
        # Example: "delivery is scheduled for two months, with payment terms requiring a 50% advance..."
        payment_description = payment_text.replace('Option ' + str(payment_option) + ': ', '')
        
        if form_type == "Bet-chem":
            # Bet-chem: Terms header in C32, payment terms directly below in C33
            # Calculate dynamic position based on total_row (which is already adjusted for products)
            # Ensure terms are always at least 5 rows after the total to avoid overlap
            min_spacing = 5
            terms_header_row = max(payment_terms_header_row, total_row + min_spacing)
            payment_terms_row = terms_header_row + 1  # Directly under header
            
            # Set "Terms and Conditions" header
            header_cell = f"{payment_terms_col}{terms_header_row}"
            ws[header_cell] = "Terms and Conditions"
            
            # Set payment terms directly under header
            payment_cell = f"{payment_terms_col}{payment_terms_row}"
            ws[payment_cell] = f"Delivery is scheduled for two months, with payment terms requiring {payment_description.lower()}"
        elif form_type == "Nyumb-Chem":
            # Nyumb-Chem: A25 has "Term of payment" label, B25 has the payment terms value
            # Set payment terms in B25
            payment_cell = f"{payment_terms_col}{payment_terms_row}"
            ws[payment_cell] = payment_description
        else:
            # Baracoda: Terms header in B28, payment terms in B30
            # Calculate dynamic position based on total_row (which is already adjusted for products)
            # Ensure terms are always at least 2 rows after the total to avoid overlap
            min_spacing = 2
            terms_header_row = max(payment_terms_header_row, total_row + min_spacing)
            payment_terms_row = terms_header_row + 2  # B30 (2 rows below header)
            
            # Set "TRADE TERMS & CONDITIONS" header
            header_cell = f"{payment_terms_col}{terms_header_row}"
            ws[header_cell] = "TRADE TERMS & CONDITIONS"
            
            # Set payment terms
            payment_cell = f"{payment_terms_col}{payment_terms_row}"
            ws[payment_cell] = f"2. Payment: {payment_description}"

    # ===== STEP 7: Save output =====
    if output_path is None:
        # Create temporary file
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, f"quotation_{form_type}_{os.getpid()}.xlsx")
    
    # Ensure the workbook has content before saving
    if ws.max_row == 0 and ws.max_column == 0:
        raise ValueError("Workbook appears to be empty. Template may not have loaded correctly.")
    
    # Save the workbook
    try:
        wb.save(output_path)
        # Verify the file was created and has content
        if not os.path.exists(output_path):
            raise IOError(f"Failed to create output file at: {output_path}")
        file_size = os.path.getsize(output_path)
        if file_size == 0:
            raise IOError(f"Output file is empty (0 bytes) at: {output_path}")
    except Exception as e:
        raise IOError(f"Failed to save quotation file: {str(e)}")
    
    return output_path


def get_template_path(form_type: str = "Baracoda") -> str:
    """
    Get the template path for a given form type.
    
    Args:
        form_type: Type of quotation form (Baracoda, Nyumb-Chem, Bet-chem)
    
    Returns:
        Path to the template file
    """
    # Get the project root directory
    # __file__ is backend/app/services/quotation_service.py
    # parent.parent.parent.parent goes: services -> app -> backend -> project root
    current_dir = Path(__file__).parent.parent.parent.parent
    
    # Select template filename based on form type
    template_filename = "PFI Sample Baracodda (2).xlsx"
    if form_type == "Nyumb-Chem":
        template_filename = "nyumbchem PFI Template.xlsx"
    elif form_type == "Bet-chem":
        template_filename = "betchem pfi TEMPLATE.xlsx"
    
    # Build template path and resolve to absolute path
    template_path = current_dir / "qoute_format" / template_filename
    template_path = template_path.resolve()
    
    if not template_path.exists():
        # Try alternative paths
        alt_paths = []
        
        # Try relative to current working directory
        alt_path1 = Path("qoute_format") / template_filename
        alt_paths.append(alt_path1)
        
        # Try from backend directory
        alt_path2 = Path(__file__).parent.parent.parent / "qoute_format" / template_filename
        alt_paths.append(alt_path2)
        
        # Try from project root (one level up from backend)
        alt_path3 = Path(__file__).parent.parent.parent.parent / "qoute_format" / template_filename
        alt_paths.append(alt_path3)
        
        # Check each alternative path
        for alt_path in alt_paths:
            if alt_path.exists():
                return str(alt_path.resolve())
        
        # If none found, raise error with all tried paths
        tried_paths = [str(template_path)] + [str(p.resolve()) if p.exists() else f"NOT FOUND: {p}" for p in alt_paths]
        raise FileNotFoundError(
            f"Template file not found for form_type '{form_type}'\n"
            f"Looking for: {template_filename}\n"
            f"Current working directory: {os.getcwd()}\n"
            f"Service file location: {Path(__file__).resolve()}\n"
            f"Tried paths:\n" + "\n".join(f"  {i+1}. {path}" for i, path in enumerate(tried_paths)) + "\n"
            f"Please ensure the template file exists in the qoute_format folder."
        )
    
    return str(template_path)

