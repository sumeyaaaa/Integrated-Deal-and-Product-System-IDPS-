"""
CRM Service - Business Logic Layer

This file contains the "business logic" for CRM operations.
Think of it as the "middle layer" between:
- API routes (what users call)
- Database (where data is stored)

Why separate? 
- Keeps API routes clean and simple
- Makes business logic reusable
- Easier to test
"""
from typing import List, Optional, Dict, Any
from datetime import datetime
import json
import re
import tempfile
import logging
from pathlib import Path

from supabase import Client
from thefuzz import fuzz
from openpyxl import load_workbook

from app.database.connection import get_supabase_client
from app.models.crm import (
    Customer,
    CustomerCreate,
    CustomerUpdate,
    Interaction,
    InteractionCreate,
    InteractionUpdate,
    QuoteDraftRequest,
    DashboardMetrics,
)
from app.services.ai_service import gemini_chat, gemini_embed, log_conversation_to_rag, search_documents

# Sales stage definitions (Brian Tracy 7-stage process)
SALES_STAGES = {
    "1": "Prospecting",
    "2": "Rapport",
    "3": "Needs Analysis",
    "4": "Presenting Solution",
    "5": "Handling Objections",
    "6": "Closing",
    "7": "Follow-up & Cross-sell",
}
from app.services.web_search_service import search_web_for_company, search_linkedin_profiles_ethiopia
from app.services.pms_service import get_all_categories


# =============================
# CUSTOMER SERVICES
# =============================


def get_all_customers(
    limit: int = 100,
    offset: int = 0,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[Customer]:
    """Get all customers from the database with pagination.
    
    If date filters are provided, only returns customers that have interactions
    within the specified date range.
    
    Args:
        limit: Maximum number of customers to return
        offset: Number of customers to skip
        start_date: Optional ISO date string (YYYY-MM-DD) - filter customers with interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter customers with interactions up to this date
    """
    supabase: Client = get_supabase_client()

    # If date filters are provided, get customers that have interactions in that range
    if start_date or end_date:
        # First, get distinct customer_ids from interactions table with date filter
        interaction_query = supabase.table("interactions").select("customer_id")
        
        if start_date:
            interaction_query = interaction_query.gte("created_at", f"{start_date}T00:00:00")
        if end_date:
            interaction_query = interaction_query.lte("created_at", f"{end_date}T23:59:59")
        
        interactions_res = interaction_query.execute()
        customer_ids = list(set(row.get("customer_id") for row in (interactions_res.data or [])))
        
        if not customer_ids:
            return []
        
        # Then fetch those customers
        response = (
            supabase.table("customers")
            .select("*")
            .in_("customer_id", customer_ids)
            .order("created_at", desc=True)
            .limit(limit)
            .offset(offset)
            .execute()
        )
    else:
        # Normal query without date filtering
        response = (
            supabase.table("customers")
            .select("*")
            .order("created_at", desc=True)
            .limit(limit)
            .offset(offset)
            .execute()
    )

    return [Customer(**row) for row in (response.data or [])]


def get_customer_by_id(customer_id: str) -> Optional[Customer]:
    """Get a single customer by UUID."""
    supabase: Client = get_supabase_client()

    response = (
        supabase.table("customers")
        .select("*")
        .eq("customer_id", customer_id)
        .execute()
    )

    if response.data:
        return Customer(**response.data[0])
    return None


def get_customers_count() -> int:
    """Get total number of customers in the database."""
    supabase: Client = get_supabase_client()

    response = supabase.table("customers").select("customer_id", count="exact").execute()

    return response.count if getattr(response, "count", None) is not None else 0


def search_customers_by_name(query: str, limit: int = 20) -> List[Customer]:
    """Search customers by partial name (case-insensitive).

    This uses Postgres ILIKE under the hood to match anywhere in the name.
    Example: query='sika' will match 'Sika Abyssinia', 'Sika Ethiopia', etc.
    """
    supabase: Client = get_supabase_client()

    response = (
        supabase.table("customers")
        .select("*")
        .ilike("customer_name", f"%{query}%")
        .order("customer_name", desc=False)
        .limit(limit)
        .execute()
    )

    return [Customer(**row) for row in (response.data or [])]


def _generate_display_id() -> str:
    """Generate a human-readable customer ID like LC-YYYY-CUST-0001.

    This mirrors the logic from the original Streamlit app so that
    new customers in the FastAPI backend keep a similar ID format.
    """
    supabase: Client = get_supabase_client()
    year = datetime.now().year

    # Fetch existing display IDs to find the highest counter for this year
    response = supabase.table("customers").select("display_id").execute()

    max_num = 0
    for row in response.data or []:
        display_id = row.get("display_id") or ""
        prefix = f"LC-{year}-CUST-"
        if isinstance(display_id, str) and display_id.startswith(prefix):
            try:
                num = int(display_id.split("-")[-1])
                if num > max_num:
                    max_num = num
            except ValueError:
                continue

    new_num = max_num + 1
    return f"LC-{year}-CUST-{new_num:04d}"


def create_customer(customer_in: CustomerCreate) -> Customer:
    """Create a new customer.

    If display_id is not provided, a new one is generated automatically.
    """
    supabase: Client = get_supabase_client()

    # ---------------------------------------------
    # 1) Duplicate-check inspired by Streamlit v6
    # ---------------------------------------------
    # Look for customers with similar names using a fuzzy match.
    # In the original Streamlit app this was a separate step in the UI;
    # here we surface it as a clear error so the frontend can warn the user.
    existing_response = (
        supabase.table("customers")
        .select("customer_id, customer_name, display_id")
        .ilike("customer_name", f"%{customer_in.customer_name}%")
        .limit(20)
        .execute()
    )
    similar_names: List[str] = []
    for row in existing_response.data or []:
        name = (row.get("customer_name") or "").strip()
        if not name:
            continue
        score = fuzz.partial_ratio(
            customer_in.customer_name.lower(),
            name.lower(),
        )
        if score >= 85:
            display_id = row.get("display_id") or "—"
            similar_names.append(f"{name} (ID: {display_id}, score: {score})")

    if similar_names:
        # Let the API layer translate this into a 409 Conflict.
        joined = "; ".join(similar_names[:3])
        raise ValueError(
            f"Similar customers already exist. Please review before creating a new one: {joined}"
        )

    # ---------------------------------------------
    # 2) Create the base customer row
    # ---------------------------------------------
    data = customer_in.model_dump()
    if not data.get("display_id"):
        data["display_id"] = _generate_display_id()

    response = supabase.table("customers").insert(data).execute()

    if not response.data:
        raise RuntimeError("Failed to create customer")

    customer = Customer(**response.data[0])
    return customer


def update_customer(customer_id: str, customer_update: CustomerUpdate) -> Customer:
    """Update an existing customer."""
    
    supabase: Client = get_supabase_client()
    
    # Check if customer exists
    existing = get_customer_by_id(customer_id)
    if not existing:
        raise ValueError("Customer not found")
    
    # Build update payload (only include fields that are provided)
    update_data = customer_update.model_dump(exclude_unset=True)
    
    if not update_data:
        # No fields to update
        return existing
    
    # If updating customer_name, check for duplicates
    if "customer_name" in update_data:
        existing_response = (
            supabase.table("customers")
            .select("customer_id, customer_name, display_id")
            .ilike("customer_name", f"%{update_data['customer_name']}%")
            .limit(20)
            .execute()
        )
        for row in existing_response.data or []:
            if str(row.get("customer_id")) != customer_id:
                name = (row.get("customer_name") or "").strip()
                if name:
                    score = fuzz.partial_ratio(
                        update_data["customer_name"].lower(),
                        name.lower(),
                    )
                    if score >= 85:
                        raise ValueError(f"Similar customer already exists: {name}")
    
    # Update the customer
    response = (
        supabase.table("customers")
        .update(update_data)
        .eq("customer_id", customer_id)
        .execute()
    )
    
    if not response.data:
        raise RuntimeError("Failed to update customer")

    return Customer(**response.data[0])


def delete_customer(customer_id: str) -> None:
    """Delete a customer and all associated interactions (cascade)."""
    supabase: Client = get_supabase_client()
    
    # Check if customer exists
    existing = get_customer_by_id(customer_id)
    if not existing:
        raise ValueError("Customer not found")
    
    # Delete customer (interactions will be cascade deleted by database foreign key)
    supabase.table("customers").delete().eq("customer_id", customer_id).execute()


def build_customer_profile(customer_id: str, user_id: Optional[str] = None) -> Customer:
    """
    Generate a customer profile using AI, existing conversations, and web search.
    
    This function replicates the exact logic from the Streamlit MVP (v6-gemini-final.py):
    - Searches relevant documents and memories (RAG)
    - Searches web for company information (Google PSE, SerpAPI, Wikipedia)
    - Searches LinkedIn for decision-makers
    - Generates comprehensive profile with Strategic-Fit Matrix
    - Uses dynamic product categories from chemical_types table
    
    This is called on-demand when the user clicks "Build Profile" button.
    """
    supabase: Client = get_supabase_client()
    
    # Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise RuntimeError("Customer not found")
    
    # Step 1: Search relevant documents and memories (RAG)
    try:
        relevant_docs = search_documents(customer.customer_name, user_id=user_id, limit=3)
    except Exception as e:
        logging.warning(f"Document search failed: {str(e)}")
        relevant_docs = []
    
    # Step 2: Search web for company information
    try:
        web_context = search_web_for_company(customer.customer_name)
    except Exception as e:
        logging.warning(f"Web search failed: {str(e)}")
        web_context = ""
    
    # Step 3: Search for LinkedIn profiles in Ethiopia
    try:
        linkedin_context = search_linkedin_profiles_ethiopia(customer.customer_name)
    except Exception as e:
        logging.warning(f"LinkedIn search failed: {str(e)}")
        linkedin_context = ""
    
    # Step 4: Combine all context
    context = ""
    if relevant_docs:
        context += "\nRelevant conversations:\n"
        for doc in relevant_docs:
            context += f"\n{doc.get('content', '')}\n"
    
    if web_context:
        context += "\nWeb Search Results:\n"
        context += web_context
    
    if linkedin_context:
        context += "\nLinkedIn Information:\n"
        context += linkedin_context
    
    # Step 5: Fetch unique categories from chemical_types table (dynamically)
    try:
        categories_list = get_all_categories()
        if not categories_list:
            categories_list = ["Cement", "Dry-Mix", "Admixtures", "Paint & Coatings"]
    except Exception as e:
        logging.warning(f"Failed to fetch categories: {str(e)}. Using defaults.")
        categories_list = ["Cement", "Dry-Mix", "Admixtures", "Paint & Coatings"]
    
    # Build category list for prompt (normalize keys for JSON: lowercase, underscores)
    category_prompt_lines = []
    category_json_keys = {}
    for cat in categories_list:
        json_key = cat.lower().replace(" ", "_").replace("-", "_").replace("&", "and")
        category_json_keys[cat] = json_key
        category_prompt_lines.append(f"- {cat} (0=No Fit, 1=Low Fit, 2=Moderate Fit, 3=High Fit)")
    
    categories_text = "\n".join(category_prompt_lines)
    json_example_keys = {json_key: "0-3" for json_key in category_json_keys.values()}
    json_example = json.dumps({"strategic_fit_matrix": json_example_keys}, indent=2)
    
    # Step 6: Create the enhanced system prompt (based on Streamlit MVP, simplified for readability)
    # NOTE: We explicitly control style so the output is clean and easy to read inside the CRM UI.
    system_prompt = f"""You are an Industry-Intel Research Assistant and B2B Chemical-Supply Strategist for LeanChem. Your mission is to perform a deep-dive analysis of the target company and all of its construction-relevant subsidiaries operating in Ethiopia, to:

Identify all business units manufacturing products for the construction sector: cement, dry-mix mortar, concrete admixtures, and paint/coatings.

Evaluate how LeanChem's chemical portfolio aligns with each unit's product and operational profile.

Recommend precise engagement strategies tailored by subsector and supply pain points.

Provide verified decision-maker contacts for B2B outreach.

If the company is a conglomerate, list all major business units and subsidiaries relevant to construction, chemicals, and manufacturing, even if not all are found in the immediate context.

🧾 Primary Deliverables
Company Overview & Recent News

≤500-character summary of the target company's core business, size, and activity in Ethiopia.

Highlight recent expansions, investments, or new product lines in cement, dry-mix, admixtures, or coatings, using GPT-4o/web-search insights or official sources.

Include citations [1], [2], … from reliable sources.

Construction-Sector Manufacturing Overview

Give a SHORT, readable list (not a table) of business units in Ethiopia manufacturing construction-related materials.
For each relevant unit, show on a single line:
- Business Unit – Construction Products – Location (City, Country) – Scale Metric (optional) – Source

Strategic-Fit Matrix

For each relevant subsidiary, assess alignment to LeanChem's offerings across {len(categories_list)} subsectors:
{categories_text}

Score each axis using:

0 = No Fit

1 = Low Fit

2 = Moderate Fit

3 = High Fit

Base scores on:

Volume opportunity vs LeanChem capacity

LeanChem's ability to solve supply or technical pain points (e.g., forex, lead time, performance)

Competitive pressure and likelihood of switching

Strategic Insights & Action Plan

Max 150-word narrative outlining 3–5 high-leverage opportunities and pain-point matches.

Segment by subsector ({', '.join(categories_list)}) and recommend clear engagement actions such as:
- Outreach channel (email, event, enabler)
- Sample trial with product match
- Proposal for supply contract, JIT, or SEZ warehousing
- Technical advisory to improve performance or reduce cost

Key Contacts for Engagement

List up to 10 decision-makers or influencers in operations, procurement, or technical roles.

Columns:

Name

Position

LinkedIn Profile (full clickable URL)

Source

Extract only real individuals verified via LinkedIn or company websites.

 Research Inputs
LeanChem Offerings

Dry-Mix/Plaster: RDP, HPMC, Starch Ether, Fiber, Zinc Stearate, Plasticizer, Defoamer, SBR, Acrylic Waterproofing, White Cement, Iron Oxide, Titanium Dioxide

Concrete Admixtures: PCE, SNF, Lignosulphonate, Sodium Gluconate, Penetrol-type waterproofing

Paint/Coatings: Styrene-Acrylic Binders, Pure Acrylics, VAE, HEC, White Cement, Iron Oxide, Titanium Dioxide

Cement Grinding: Cement grinding aids

🔍 Research Tools & Constraints
Source from:

The target company's official website and group/subsidiary pages

Annual reports and press releases

LinkedIn (for verified role-based contacts)

News outlets, trade journals, government registries

Use structured search queries like:
- "[Target Company] cement plant Ethiopia"
- "[Target Company] paint coatings manufacturer Ethiopia"
- "[Target Company] dry mix mortar factory site"
- "[Target Company] procurement manager LinkedIn"

Use numbered citations [1], [2], etc. only when they truly help.

Provide honest results—if a construction vertical is not present, list as "N/A" or "0" in the fit matrix.

STYLE REQUIREMENTS (CRITICAL - FOLLOW EXACTLY):
- ABSOLUTELY NO MARKDOWN: No tables (| ... |), no asterisks (* or **), no code fences (```), no emojis, no markdown links [text](url).
- Use ONLY plain text with simple line breaks.
- Use exactly 4 numbered sections: "1. Company Snapshot", "2. Construction Footprint in Ethiopia", "3. Strategic Fit Assessment", "4. Recommended Next Steps".
- For Strategic-Fit Matrix: Write one simple line per category like "Admixtures: 0/3 - No manufacturing presence in Ethiopia" or "Paint & Coatings: 2/3 - Potential for raw material supply".
- For business units: Use simple bullet points with dashes, one per line, like "- Unit Name: Products - Location - Scale".
- For contacts: List as simple lines like "Name: [Name], Position: [Title], LinkedIn: [URL]".
- Keep sentences short. Maximum 2-3 sentences per paragraph.
- Remove all citations [1], [2] from the main text - only mention sources naturally in sentences if needed.
- Total length should be under 800 words, easy to scan in 2 minutes.

CRITICAL: At the END of your response, include a JSON block with the Strategic-Fit Matrix scores:
{json_example}

Use the exact category names as keys (lowercase, underscores for spaces)."""
    
    # Step 7: Build messages with context
    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": f"Generate a profile for: {customer.customer_name}\n\nContext:{context}"
        }
    ]
    
    # Step 8: Get AI response
    try:
        profile_text = gemini_chat(messages)
        if not profile_text or not profile_text.strip():
            raise RuntimeError("AI service returned empty response. Please check GEMINI_API_KEY configuration.")
    except Exception as e:
        error_msg = f"Failed to generate AI profile: {str(e)}"
        logging.error(error_msg)
        raise RuntimeError(error_msg)
    
    # Step 8.5: Post-process to remove any markdown that slipped through
    # Remove markdown tables (aggressive cleanup)
    # Remove table blocks - find multi-line tables
    table_pattern = r'\n\|[^\n]*\|(?:\n\|[^\n]*\|)*'
    profile_text = re.sub(table_pattern, '', profile_text)
    # Remove any remaining single-line table rows
    profile_text = re.sub(r'\|[^\n]*\|', lambda m: m.group(0).replace('|', ' • ').strip(' • '), profile_text)
    # Remove markdown formatting
    profile_text = re.sub(r'\*\*([^*]+)\*\*', r'\1', profile_text)
    profile_text = re.sub(r'\*([^*]+)\*', r'\1', profile_text)
    profile_text = re.sub(r'\[([^\]]+)\]\([^)]+\)', r'\1', profile_text)
    profile_text = re.sub(r'\[(\d+)\]', '', profile_text)
    profile_text = re.sub(r'^#{1,6}\s+', '', profile_text, flags=re.MULTILINE)
    
    # Parse the Strategic-Fit Matrix from the AI response
    json_match = None
    json_patterns = [
        r'\{[^{}]*"strategic_fit_matrix"[^{}]*\{[^{}]*\}[^{}]*\}',  # Nested
        r'\{[^}]*"strategic_fit_matrix"[^}]*\}',  # Simple
    ]
    for pattern in json_patterns:
        json_match = re.search(pattern, profile_text, re.IGNORECASE | re.DOTALL)
        if json_match:
            break
    
    # If no match, try to find any JSON block at the end of the response
    if not json_match:
        json_candidates = re.findall(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', profile_text, re.IGNORECASE | re.DOTALL)
        for candidate in reversed(json_candidates):
            if "strategic_fit_matrix" in candidate.lower():
                json_match = re.search(re.escape(candidate), profile_text, re.IGNORECASE | re.DOTALL)
                break
    
    if json_match:
        try:
            json_str = json_match.group(0)
            parsed = json.loads(json_str)
            matrix = parsed.get("strategic_fit_matrix", {})
            
            # Build product_scores dict using the actual category names as keys
            product_scores = {}
            for cat in categories_list:
                json_key = category_json_keys[cat]
                # Try both the JSON key and the original category name
                score = matrix.get(json_key) or matrix.get(cat.lower()) or matrix.get(cat)
                if score is None:
                    score = 0
                try:
                    product_scores[cat] = max(0, min(3, int(score)))
                except (ValueError, TypeError):
                    product_scores[cat] = 0
        except (json.JSONDecodeError, ValueError, KeyError):
            # If parsing fails, set defaults for all categories
            product_scores = {cat: 0 for cat in categories_list}
    else:
        # If no JSON found, set defaults for all categories
        product_scores = {cat: 0 for cat in categories_list}
    
    # Update the customer record with product alignment scores
    if product_scores:
        supabase.table("customers").update({
            "product_alignment_scores": product_scores
        }).eq("customer_id", customer.customer_id).execute()

    # Store as an interaction so the history view shows it.
    interaction_payload = InteractionCreate(
        input_text=f"System: AI profile generated for {customer.customer_name}",
        ai_response=profile_text,
        tds_id=None,
    )
    _ = create_interaction(
        customer_id=str(customer.customer_id),
        interaction_in=interaction_payload,
        user_id=user_id,
    )

    # Log into the RAG `conversation` table as a combined entry.
    try:
        combined_text = (
            f"Customer: {customer.customer_name}\n"
            f"Display ID: {customer.display_id or '—'}\n"
            f"AI-generated CRM profile:\n{profile_text}"
        )
        embedding = gemini_embed(combined_text)
        metadata = {
            "customer_id": str(customer.customer_id),
            "customer_name": customer.customer_name,
            "source": "customer_profile",
            "tds_id": None,
            "user_id": user_id,
        }
        log_conversation_to_rag(
            combined_text,
            embedding=embedding,
            metadata=metadata,
        )
    except Exception:
        # Don't block profile building if RAG logging fails.
        pass

    # Return updated customer with scores
    updated_response = supabase.table("customers").select("*").eq("customer_id", customer.customer_id).execute()
    if updated_response.data:
        return Customer(**updated_response.data[0])
    return customer


# =============================
# INTERACTION SERVICES
# =============================


def get_interactions_for_customer(
    customer_id: str,
    limit: int = 100,
    offset: int = 0,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[Interaction]:
    """Get interactions for a specific customer, newest first.
    
    Args:
        customer_id: Customer UUID
        limit: Maximum number of interactions to return
        offset: Number of interactions to skip
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    """
    supabase: Client = get_supabase_client()

    query = (
        supabase.table("interactions")
        .select("*")
        .eq("customer_id", customer_id)
    )

    # Apply date filters if provided
    if start_date:
        query = query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        query = query.lte("created_at", f"{end_date}T23:59:59")

    response = (
        query.order("created_at", desc=True)
        .limit(limit)
        .offset(offset)
        .execute()
    )

    return [Interaction(**row) for row in (response.data or [])]


def get_interactions_count_for_customer(
    customer_id: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> int:
    """Get total interaction count for a customer (for pagination).
    
    Args:
        customer_id: Customer UUID
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    """
    supabase: Client = get_supabase_client()

    query = (
        supabase.table("interactions")
        .select("id", count="exact")
        .eq("customer_id", customer_id)
    )

    # Apply date filters if provided
    if start_date:
        query = query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        query = query.lte("created_at", f"{end_date}T23:59:59")

    response = query.execute()

    return response.count if getattr(response, "count", None) is not None else 0


def get_interaction_by_id(interaction_id: str) -> Optional[Interaction]:
    """Fetch a single interaction by its UUID."""
    supabase: Client = get_supabase_client()

    response = (
        supabase.table("interactions")
        .select("*")
        .eq("id", interaction_id)
        .execute()
    )

    if response.data:
        return Interaction(**response.data[0])
    return None


def create_interaction(
    customer_id: str, interaction_in: InteractionCreate, user_id: Optional[str] = None
) -> Interaction:
    """Create a new interaction linked to a customer (and optional user)."""
    supabase: Client = get_supabase_client()

    payload = interaction_in.model_dump(exclude_unset=True)
    payload["customer_id"] = customer_id
    if user_id:
        payload["user_id"] = user_id

    response = supabase.table("interactions").insert(payload).execute()

    if not response.data:
        raise RuntimeError("Failed to create interaction")

    return Interaction(**response.data[0])


def update_interaction(
    interaction_id: str, interaction_in: InteractionUpdate
) -> Optional[Interaction]:
    """Update an existing interaction and return the updated record."""
    supabase: Client = get_supabase_client()

    update_data = interaction_in.model_dump(exclude_unset=True)
    if not update_data:
        # Nothing to update; just return the current record
        return get_interaction_by_id(interaction_id)

    response = (
        supabase.table("interactions")
        .update(update_data)
        .eq("id", interaction_id)
        .execute()
    )

    if response.data:
        return Interaction(**response.data[0])
    return None


def delete_interaction(interaction_id: str) -> None:
    """Delete an interaction. Raises on failure."""
    supabase: Client = get_supabase_client()

    supabase.table("interactions").delete().eq("id", interaction_id).execute()


def auto_fill_sales_stage_for_customer(customer_id: str) -> Optional[str]:
    """
    Analyze and set sales stage for a single customer based on their interaction history.
    
    Args:
        customer_id: The customer ID to analyze
        
    Returns:
        The new sales stage (1-7) if successful, None if customer not found or has no interactions
    """
    customer = get_customer_by_id(customer_id)
    if not customer:
        return None
    
    # If customer already has a sales stage, don't overwrite (user can manually edit)
    if customer.sales_stage:
        return customer.sales_stage
    
    try:
        # Get all interactions for this customer
        interactions = get_interactions_for_customer(customer_id, limit=50, offset=0)
        
        if not interactions:
            # No interactions = Stage 1 (Prospecting)
            supabase: Client = get_supabase_client()
            supabase.table("customers").update({"sales_stage": "1"}).eq("customer_id", customer_id).execute()
            return "1"
        
        # Build context from interactions
        history_lines = []
        for it in interactions:
            user_part = (it.input_text or "").strip()
            ai_part = (it.ai_response or "").strip()
            if user_part or ai_part:
                history_lines.append(f"Q: {user_part}\nA: {ai_part}")
        
        past_context = "\n\n".join(history_lines) if history_lines else "No past interactions"
        
        # Use the most recent interaction for analysis
        latest = interactions[0]  # Already sorted newest first
        new_interaction = f"Q: {latest.input_text or ''}\nA: {latest.ai_response or ''}"
        
        # Analyze stage
        new_stage = analyze_sales_stage(new_interaction, past_context, current_stage=None)
        
        # Update customer
        supabase: Client = get_supabase_client()
        supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
        
        return new_stage
        
    except Exception as e:
        print(f"Error auto-filling sales stage for customer {customer_id}: {e}")
        return None


def backfill_sales_stages_for_all_customers() -> Dict[str, Any]:
    """
    Analyze and set sales stages for all customers that don't have one yet.
    Uses their interaction history to determine the current stage.
    
    Returns:
        Dict with counts: {"updated": N, "skipped": M, "errors": K}
    """
    supabase: Client = get_supabase_client()
    
    # Get all customers - we'll filter for null/empty sales_stage in Python
    response = (
        supabase.table("customers")
        .select("customer_id, customer_name, sales_stage")
        .execute()
    )
    
    all_customers = response.data or []
    # Filter for customers without sales_stage
    customers_to_update = [
        c for c in all_customers 
        if not c.get("sales_stage") or c.get("sales_stage") == ""
    ]
    
    results = {"updated": 0, "skipped": 0, "errors": 0}
    
    for customer_row in customers_to_update:
        customer_id = customer_row["customer_id"]
        customer_name = customer_row["customer_name"]
        
        try:
            # Get all interactions for this customer
            interactions = get_interactions_for_customer(customer_id, limit=50, offset=0)
            
            if not interactions:
                # No interactions = Stage 1 (Prospecting)
                supabase.table("customers").update({"sales_stage": "1"}).eq("customer_id", customer_id).execute()
                results["updated"] += 1
                continue
            
            # Build context from interactions
            history_lines = []
            for it in interactions:
                user_part = (it.input_text or "").strip()
                ai_part = (it.ai_response or "").strip()
                if user_part or ai_part:
                    history_lines.append(f"Q: {user_part}\nA: {ai_part}")
            
            past_context = "\n\n".join(history_lines) if history_lines else "No past interactions"
            
            # Use the most recent interaction for analysis
            latest = interactions[0]  # Already sorted newest first
            new_interaction = f"Q: {latest.input_text or ''}\nA: {latest.ai_response or ''}"
            
            # Analyze stage
            new_stage = analyze_sales_stage(new_interaction, past_context, current_stage=None)
            
            # Update customer
            supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
            results["updated"] += 1
            
        except Exception as e:
            print(f"Error backfilling sales stage for customer {customer_id}: {e}")
            results["errors"] += 1
    
    return results


def get_dashboard_metrics(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> DashboardMetrics:
    """
    Get dashboard metrics including customer counts, interaction counts, and sales stage distribution.
    
    Args:
        start_date: Optional ISO date string (YYYY-MM-DD) - filter interactions from this date onwards
        end_date: Optional ISO date string (YYYY-MM-DD) - filter interactions up to this date
    
    Returns:
        DashboardMetrics with all calculated metrics
    """
    supabase: Client = get_supabase_client()
    
    # Get total customers count
    total_customers_response = supabase.table("customers").select("customer_id", count="exact").execute()
    total_customers = total_customers_response.count if getattr(total_customers_response, "count", None) is not None else 0
    
    # Get interactions count with date filtering
    interactions_query = supabase.table("interactions").select("id", count="exact")
    if start_date:
        interactions_query = interactions_query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        interactions_query = interactions_query.lte("created_at", f"{end_date}T23:59:59")
    
    interactions_response = interactions_query.execute()
    total_interactions = interactions_response.count if getattr(interactions_response, "count", None) is not None else 0
    
    # Get distinct customers with interactions (within date range if specified)
    customers_with_interactions_query = supabase.table("interactions").select("customer_id")
    if start_date:
        customers_with_interactions_query = customers_with_interactions_query.gte("created_at", f"{start_date}T00:00:00")
    if end_date:
        customers_with_interactions_query = customers_with_interactions_query.lte("created_at", f"{end_date}T23:59:59")
    
    customers_with_interactions_response = customers_with_interactions_query.execute()
    unique_customer_ids = set(row.get("customer_id") for row in (customers_with_interactions_response.data or []))
    customers_with_interactions = len(unique_customer_ids)
    
    # Get sales stages distribution
    customers_query = supabase.table("customers").select("sales_stage")
    customers_response = customers_query.execute()
    
    sales_stages_distribution: Dict[str, int] = {
        "1": 0, "2": 0, "3": 0, "4": 0, "5": 0, "6": 0, "7": 0
    }
    
    for row in customers_response.data or []:
        stage = row.get("sales_stage")
        if stage and stage in sales_stages_distribution:
            sales_stages_distribution[stage] = sales_stages_distribution.get(stage, 0) + 1
    
    return DashboardMetrics(
        total_customers=total_customers,
        total_interactions=total_interactions,
        customers_with_interactions=customers_with_interactions,
        sales_stages_distribution=sales_stages_distribution,
    )


def analyze_sales_stage(
    new_interaction: str,
    past_context: str,
    current_stage: Optional[str] = None,
) -> str:
    """
    Analyze the sales stage based on Brian Tracy's 7-stage process.
    
    Returns the current stage number (1-7) as a string.
    """
    system_prompt = f"""You are "LeanChem 7-Stage Sales Tracker" based on Brian Tracy's sales process.

Analyze the customer interaction and determine which of the 7 stages they are currently in:

1. Prospecting - Customer identified, initial contact made
2. Rapport - Trust built, relationship established
3. Needs Analysis - Customer shares requirements, pain points identified
4. Presenting Solution - Product/service proposal presented
5. Handling Objections - Addressing concerns, negotiating terms
6. Closing - Finalizing deal, contract signed
7. Follow-up & Cross-sell - Post-sale support, upselling opportunities

CURRENT STAGE (if known): {current_stage or "None"}

PAST CONTEXT:
\"\"\"{past_context}\"\"\"

NEW INTERACTION:
\"\"\"{new_interaction}\"\"\"

TASK:
- Analyze the NEW INTERACTION in context of PAST CONTEXT
- Determine the CURRENT stage (1-7) based on evidence
- If CURRENT STAGE is provided, only advance if there's strong new evidence
- Return ONLY the stage number (1, 2, 3, 4, 5, 6, or 7) - nothing else

OUTPUT FORMAT:
Return only a single digit: 1, 2, 3, 4, 5, 6, or 7
"""

    messages = [
        {"role": "system", "content": system_prompt},
        {"role": "user", "content": "What is the current sales stage? Return only the number (1-7)."},
    ]

    try:
        response = gemini_chat(messages).strip()
        # Extract just the number if there's extra text
        import re
        match = re.search(r'\b([1-7])\b', response)
        if match:
            stage_num = match.group(1)
            # Validate it's 1-7
            if stage_num in SALES_STAGES:
                return stage_num
        # Default to stage 1 if parsing fails
        return "1"
    except Exception:
        # Default to stage 1 on error
        return "1"


def chat_with_customer(
    customer_id: str,
    input_text: str,
    tds_id: Optional[str] = None,
    user_id: Optional[str] = None,
    file_url: Optional[str] = None,
    file_type: Optional[str] = None,
    file_content: Optional[str] = None,
) -> Interaction:
    """
    Run an AI chat turn for a specific customer:
    - Calls Gemini to generate a response
    - Stores the turn in `interactions` table
    - Logs a combined Q/A entry into `conversation` (RAG) with embedding
    """
    supabase: Client = get_supabase_client()

    # 1) Ensure customer exists
    customer = get_customer_by_id(customer_id)
    if not customer:
        raise RuntimeError("Customer not found")

    # 2) Fetch recent interactions to give the AI richer CRM context
    recent_interactions = get_interactions_for_customer(
        customer_id, limit=10, offset=0
    )
    # Oldest first in the prompt so the story reads naturally
    recent_interactions = list(reversed(recent_interactions))

    history_lines: list[str] = []
    for i, it in enumerate(recent_interactions, start=1):
        user_part = (it.input_text or "").strip()
        ai_part = (it.ai_response or "").strip()
        if not user_part and not ai_part:
            continue
        history_lines.append(
            f"Interaction {i}:\n"
            f"Input: {user_part or '[no input logged]'}\n"
            f"Output: {ai_part or '[no AI response logged]'}"
        )

    memories_str = "\n\n".join(history_lines) if history_lines else "No past interactions yet."

    customer_context = (
        f"Customer name: {customer.customer_name}\n"
        f"Display ID: {customer.display_id or '—'}\n"
        f"Customer ID: {customer.customer_id}\n"
        f"Total recorded interactions: {len(recent_interactions)}"
    )

    # 3) Prepare messages for Gemini using the same style as the Streamlit CRM chat
    system_prompt = f"""
You are a helpful AI assistant specialized in chemical trading and CRM.
If the user asks about a specific customer, use the customer's most relevant past interactions below.
Also use the provided memories and any relevant conversations from the database.
If you don't find relevant information, say so and reason transparently.

Customer context:
{customer_context}

User Memories:
{memories_str}
"""

    # 3.5) Add file content to context if provided
    user_content = input_text
    if file_content:
        user_content = f"{input_text}\n\n--- Attached File Content ---\n{file_content}"

    messages = [
        {"role": "system", "content": system_prompt},
        {
            "role": "user",
            "content": user_content,
        },
    ]

    # 4) Call Gemini
    ai_response = gemini_chat(messages)

    # 5) Store in interactions table
    interaction_payload = InteractionCreate(
        input_text=input_text,
        ai_response=ai_response,
        tds_id=tds_id,
        file_url=file_url,
        file_type=file_type,
    )
    interaction = create_interaction(customer_id, interaction_payload, user_id=user_id)

    # 6) Log to RAG `conversation` table with embedding
    try:
        combined_text = f"Customer: {customer.customer_name}\nQ: {input_text}\nA: {ai_response}"
        embedding = gemini_embed(combined_text)
        metadata = {
            "customer_id": str(customer.customer_id),
            "customer_name": customer.customer_name,
            "source": "crm_chat",
            "tds_id": tds_id,
            "user_id": user_id,
        }
        log_conversation_to_rag(combined_text, embedding=embedding, metadata=metadata)
    except Exception:
        # Don't block the main interaction flow if RAG logging fails
        pass

    # 7) Analyze and update sales stage
    try:
        # Build context for stage analysis
        stage_context = f"{customer_context}\n\nRecent Interactions:\n{memories_str}"
        new_stage = analyze_sales_stage(
            new_interaction=f"Q: {input_text}\nA: {ai_response}",
            past_context=stage_context,
            current_stage=customer.sales_stage,
        )
        
        # Update customer's sales stage if it changed
        if new_stage != customer.sales_stage:
            supabase.table("customers").update({"sales_stage": new_stage}).eq("customer_id", customer_id).execute()
    except Exception:
        # Don't block the main interaction flow if stage analysis fails
        pass

    return interaction


def generate_quote_excel(body: QuoteDraftRequest) -> str:
    """
    Generate an AI-enhanced Excel quotation file based on a template.

    Baracoda format:
      - B12 = customer name
      - C20, C21, ... = product names (starting at row 20)
      - I20, I21, ... = quantities
      - J20, J21, ... = unit prices
      - L20, L21, ... = formulas =J20*I20, =J21*I21, etc.
      - L30 = SUM(L20:L{last_row})*0.15 (VAT)
      - L31 = SUM(L20:L30) (total incl. VAT)

    Betchem format:
      - A4 = company name
      - E4 = date
      - B8, B9, ... = product names (starting at row 8)
      - C8, C9, ... = unit measurement (kg, mt, etc.)
      - D8, D9, ... = quantities
      - E8, E9, ... = unit prices before VAT
      - F8, F9, ... = formulas =D8*E8, =D9*E9, etc.
      - F12 = SUM(F8:F{last_row}) (subtotal)
      - F13 = F12*0.15 (VAT)
      - F14 = F12+F13 (total incl. VAT)
    """
    # 1) Locate template on disk
    fmt = body.format.lower()
    if fmt == "baracoda":
        filename = "Baracoda.xlsx"
    elif fmt == "betchem":
        filename = "Betchem.xlsx"
    else:
        raise RuntimeError(f"Unsupported quote format: {body.format}")

    template_path = Path(__file__).resolve().parents[3] / "qoute_format" / filename
    if not template_path.exists():
        raise RuntimeError(f"Template file not found: {template_path}")

    # 2) Open template and get the first/active sheet
    wb = load_workbook(template_path)
    ws = wb.active

    # 3) Parse quantities and prices for all products
    parsed_products = []
    for p in body.products:
        # Parse quantity
        try:
            qty_val = float(p.quantity)
        except Exception:
            qty_val = None

        # Parse unit price
        unit_price_val: Optional[float] = None
        if p.target_price is not None:
            try:
                unit_price_val = float(p.target_price)
            except Exception:
                # Try to extract leading numeric part like "1200" from "1200 USD/MT"
                m = re.match(r"\s*([0-9]+(?:\.[0-9]+)?)", str(p.target_price))
                if m:
                    try:
                        unit_price_val = float(m.group(1))
                    except Exception:
                        unit_price_val = None

        parsed_products.append({
            "name": p.chemical_type_name,
            "unit": p.unit,
            "quantity": qty_val if qty_val is not None else p.quantity,
            "unit_price": unit_price_val if unit_price_val is not None else (p.target_price or ""),
        })

    # 4) Apply format-specific mappings
    if fmt == "baracoda":
        # Baracoda format
        ws["B12"] = body.customer_name

        start_row = 20
        last_product_row = start_row + len(parsed_products) - 1

        for idx, p in enumerate(parsed_products):
            row = start_row + idx
            ws[f"C{row}"] = p["name"]
            ws[f"I{row}"] = p["quantity"]
            ws[f"J{row}"] = p["unit_price"]
            ws[f"L{row}"] = f"=J{row}*I{row}"

        ws["L30"] = f"=SUM(L{start_row}:L{last_product_row})*0.15"
        ws["L31"] = f"=SUM(L{start_row}:L30)"

        # Write terms and conditions to B34 for Baracoda
        # Use provided terms or default
        terms_text = body.terms_and_conditions or "Terms and conditions:\nMinium Order Quantity: 1000 KG Per Product\nPayment: Advance Payment is 50% 30% when goods are delivered at Moyale and & Balance Payment is 20% on Delivery."
        cell = ws["B34"]
        cell.value = terms_text
        # Enable text wrapping for multi-line content
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    elif fmt == "betchem":
        # Betchem format
        ws["A4"] = body.customer_name
        # Write current date to E4
        from datetime import datetime
        ws["E4"] = datetime.now().strftime("%Y-%m-%d")

        start_row = 8
        last_product_row = start_row + len(parsed_products) - 1

        for idx, p in enumerate(parsed_products):
            row = start_row + idx
            ws[f"B{row}"] = p["name"]
            ws[f"C{row}"] = p["unit"]  # Unit measurement (kg, mt, etc.)
            ws[f"D{row}"] = p["quantity"]
            ws[f"E{row}"] = p["unit_price"]
            ws[f"F{row}"] = f"=D{row}*E{row}"

        # F12 = sum of all product subtotals
        ws["F12"] = f"=SUM(F{start_row}:F{last_product_row})"
        # F13 = VAT (15%)
        ws["F13"] = "=F12*0.15"
        # F14 = total including VAT
        ws["F14"] = "=F12+F13"

        # Write terms and conditions to A16 for Betchem
        # Use provided terms or default
        terms_text = body.terms_and_conditions or "Terms and conditions:\n- For items currently in stock, the advance payment is 100 %"
        cell = ws["A16"]
        cell.value = terms_text
        # Enable text wrapping for multi-line content
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 5) Save to a temporary file and return its path
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=f"_{filename}")
    tmp_path = Path(tmp.name)
    tmp.close()
    wb.save(tmp_path)

    return str(tmp_path)